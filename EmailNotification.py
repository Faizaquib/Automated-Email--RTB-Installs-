#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
import os
import re
import time
import smtplib

def sendEmail(Sender,Reciever,fileName):
    smtpServer='smtp-relay.worldpay.local'      
    fromAddr = Sender       
    toAddr = Reciever
    text = "Subject: Notification - File Available\n\n"+fileName+" is a now available.\n\n Cheers :) "
    server = smtplib.SMTP(smtpServer,25)
    server.ehlo()
    server.starttls()
    server.sendmail(fromAddr, toAddr.split(","), text) 
    server.quit()
    

def getFileName(folderPath,fileName):
    listOfFileFound = []
    filesInFolder = os.listdir(folderPath)
    for iterator in range(0,len(filesInFolder)):
        if fileName in filesInFolder[iterator]:
            listOfFileFound.append(filesInFolder[iterator])
    return listOfFileFound        
    

def ToLookFor():
    folderPathToLookFor = []
    fileNameToLookFor = []
    Sender = []
    Reciever = []
    path = r"C:\AppInstalls\AutoNotification.xlsx"
    workBook = openpyxl.load_workbook(path)
    sheet = workBook.active
    maxRow = sheet.max_row
    for iterator in range(2,maxRow+1):
        folderPathToLookFor.append(sheet.cell(row = iterator, column = 1).value)
        fileNameToLookFor.append(sheet.cell(row = iterator, column = 2).value)
        Sender.append(sheet.cell(row = iterator, column = 3).value)
        Reciever.append(sheet.cell(row = iterator, column = 4).value)
    return folderPathToLookFor,fileNameToLookFor,Sender,Reciever
 
def getSenbederReciever(fileName):
    path = r"C:\AppInstalls\AutoNotification.xlsx"
    workBook = openpyxl.load_workbook(path)
    sheet = workBook.active
    maxRow = sheet.max_row
    for iterator in range(2,maxRow+1):
        if sheet.cell(row = iterator, column = 2).value in fileName:
            Sender = sheet.cell(row = iterator, column = 3).value
            reciever = sheet.cell(row = iterator, column = 4).value
            return Sender,reciever
    
def Notification():
    alreadyProcessedFiles = []
    for counter in range(1,97):
        print(counter)
        files= []
        folderPathToLookFor,fileNameToLookFor,Sender,Reciever = ToLookFor()
        if len(folderPathToLookFor)>0:
            for iterator in range(0,len(folderPathToLookFor)):
                files = files + getFileName(folderPathToLookFor[iterator],fileNameToLookFor[iterator])
                
            for iterator in range(0,len(files)):
                if files[iterator] in alreadyProcessedFiles:
                    continue
                    
                else:
                    Sender,Reciever = getSenbederReciever(files[iterator])
                    sendEmail(Sender,Reciever,files[iterator])
                    alreadyProcessedFiles.append(files[iterator])
                    
        time.sleep(300)  
        
Notification()        

